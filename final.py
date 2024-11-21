import tkinter as tk
from tkinterdnd2 import TkinterDnD, DND_FILES
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

def browse_file(label, side):
    filename = filedialog.askopenfilename(title=f"Select the {side} Excel file", filetypes=[("Excel files", "*.xlsx")])
    if filename:
        label.config(text=filename)
    return filename

def drop_file(event, label):
    filename = event.data
    if filename.endswith('.xlsx'):
        label.config(text=filename)
    else:
        messagebox.showerror("Error", "Please drop an Excel file (.xlsx)")

def clear_file(label):
    label.config(text="Drag and drop or click to upload the file")
def compare_sheets():
    file1 = left_label.cget("text")
    file2 = right_label.cget("text")

    if not file1 or not file2 or "Drag" in file1 or "Drag" in file2:
        messagebox.showerror("Error", "Please select or drop both files before comparing.")
        return

    try:
        df1 = pd.read_excel(file1)
        df2 = pd.read_excel(file2)
        result_df = df1.copy()

        # Ask for output file location
        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_file:
            return  # User cancelled the save dialog

        # Compare the two DataFrames
        for row in range(len(df1)):
            for col in range(len(df1.columns)):
                cell_value1 = df1.iloc[row, col]
                cell_value2 = df2.iloc[row, col] if row < len(df2) and col < len(df2.columns) else None
                
                # If there is a difference, mark the cell and add a comment
                if pd.notna(cell_value1) and pd.notna(cell_value2) and cell_value1 != cell_value2:
                    result_df.iloc[row, col] = "Error"
                    comment = f"Original: {cell_value1}\nCompared: {cell_value2}"
                    # Create a new workbook to save the results
                    result_df.iloc[row, col] = "Error"  # Mark the error in the DataFrame

        # Save the result DataFrame to an Excel file
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='Comparison', index=False)

        # Load the workbook to add comments and formatting
        wb = load_workbook(output_file)
        ws = wb['Comparison']
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        # Highlight cells with errors and add comments
        for row in range(2, len(result_df) + 2):
            for col in range(1, len(result_df.columns) + 1):
                if ws.cell(row=row, column=col).value == "Error":
                    # Set the comment for the cell
                    cell_value1 = df1.iloc[row - 2, col - 1]  # Adjust for header row
                    cell_value2 = df2.iloc[row - 2, col - 1] if row - 2 < len(df2) else None
                    comment = f"Original: {cell_value1}\nCompared: {cell_value2}"
                    ws.cell(row=row, column=col).comment = Comment(comment, "Comparison Tool")
                    ws.cell(row=row, column=col).fill = red_fill  # Highlight error

        # Save the workbook after adding comments
        wb.save(output_file)

        messagebox.showinfo("Success", f"Comparison complete. Results saved to {output_file}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")



def animate_opening(app):
    def animate_resize(width):
        current_width = app.winfo_width()
        if current_width < width:
            app.geometry(f"{current_width + 10}x500")
            app.after(10, animate_resize, width)
        else:
            app.geometry(f"{width}x500")
    
    # Start the animation with a desired width
    animate_resize(900)

app = TkinterDnD.Tk()
app.title("Excel Comparator")
app.geometry("200x500")  # Start with a smaller size for animation

# Call the animation function
app.after(100, lambda: animate_opening(app))

app.configure(bg="#f5f5f5")

# Main Frame
main_frame = tk.Frame(app, bg="#f5f5f5")
main_frame.pack(fill=tk.BOTH, expand=True)

# Left Section (File to Compare)
left_frame = tk.Frame(main_frame, padx=20, pady=20, bg="#ffffff", relief=tk.RAISED, borderwidth=2)
left_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

left_label = tk.Label(left_frame, text="Drag and drop or click to upload the file to compare", bg="#f0f0f0", fg="#333333", width=40, height=10, relief=tk.SUNKEN, borderwidth=2, font=("Arial", 12))
left_label.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
left_label.drop_target_register(DND_FILES)
left_label.dnd_bind('<<Drop>>', lambda event: drop_file(event, left_label))

left_button_frame = tk.Frame(left_frame, bg="#ffffff")
left_button_frame.pack(fill=tk.X, pady=10)

left_button = tk.Button(left_button_frame, text="Upload File", command=lambda: browse_file(left_label, "first"), bg="#007bff", fg="#ffffff", font=("Arial", 10), relief=tk.FLAT)
left_button.pack(side=tk.LEFT, padx=5)

left_clear_button = tk.Button(left_button_frame, text="Clear", command=lambda: clear_file(left_label), bg="#dc3545", fg="#ffffff", font=("Arial", 10), relief=tk.FLAT)
left_clear_button.pack(side=tk.LEFT, padx=5)

# Right Section (Comparison File)
right_frame = tk.Frame(main_frame, padx=20, pady=20, bg="#ffffff", relief=tk.RAISED, borderwidth=2)
right_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")

right_label = tk.Label(right_frame, text="Drag and drop or click to upload the comparison file", bg="#f0f0f0", fg="#333333", width=40, height=10, relief=tk.SUNKEN, borderwidth=2, font=("Arial", 12))
right_label.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
right_label.drop_target_register(DND_FILES)
right_label.dnd_bind('<<Drop>>', lambda event: drop_file(event, right_label))

right_button_frame = tk.Frame(right_frame, bg="#ffffff")
right_button_frame.pack(fill=tk.X, pady=10)

right_button = tk.Button(right_button_frame, text="Upload File", command=lambda: browse_file(right_label, "second"), bg="#28a745", fg="#ffffff", font=("Arial", 10), relief=tk.FLAT)
right_button.pack(side=tk.LEFT, padx=5)

right_clear_button = tk.Button(right_button_frame, text="Clear", command=lambda: clear_file(right_label), bg="#dc3545", fg="#ffffff", font=("Arial", 10), relief=tk.FLAT)
right_clear_button.pack(side=tk.LEFT, padx=5)

# Compare Button
compare_button = tk.Button(main_frame, text="Compare Excel Sheets", command=compare_sheets, bg="#ff5722", fg="#ffffff", font=("Arial", 12), relief=tk.FLAT)
compare_button.grid(row=1, column=0, columnspan=2, pady=20)

# Configure grid weights
main_frame.grid_rowconfigure(0, weight=1)
main_frame.grid_columnconfigure(0, weight=1)
main_frame.grid_columnconfigure(1, weight=1)

app.mainloop()
